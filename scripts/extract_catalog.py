from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "North America Sales List.xlsx"
ASSET_ROOT = ROOT / "North America Sales List-FILE"
OUTPUT_PATH = ROOT / "catalog-data.js"

SKIP_SHEETS = {"README", "AD MAX"}


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def compact_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def parse_solution_refs(raw: str) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[,/，、\s]+", raw)
    return [part for part in (p.strip() for p in parts) if part]


def build_asset_index(asset_root: Path) -> tuple[dict[str, Path], dict[str, list[Path]]]:
    folder_index = {}
    for directory in asset_root.iterdir():
        if directory.is_dir():
            folder_index[compact_key(directory.name)] = directory

    file_index: dict[str, list[Path]] = defaultdict(list)
    for file_path in asset_root.rglob("*"):
        if file_path.is_file():
            file_index[file_path.name.lower()].append(file_path)

    return folder_index, file_index


def best_folder(sheet_name: str, folder_index: dict[str, Path]) -> Path | None:
    key = compact_key(sheet_name)
    if key in folder_index:
        return folder_index[key]

    best_match = None
    best_score = -1
    for folder_key, folder_path in folder_index.items():
        score = 0
        if key and (key in folder_key or folder_key in key):
            score += min(len(key), len(folder_key))
        common_prefix = 0
        for left, right in zip(key, folder_key):
            if left != right:
                break
            common_prefix += 1
        score += common_prefix
        if score > best_score:
            best_score = score
            best_match = folder_path
    return best_match


def resolve_assets(image_cell: str, sheet_folder: Path | None, file_index: dict[str, list[Path]]) -> list[str]:
    if not image_cell:
        return []

    results: list[str] = []
    wanted = [part.strip() for part in image_cell.split(",") if part.strip()]
    for asset_name in wanted:
        candidates = file_index.get(asset_name.lower(), [])
        picked = None
        if sheet_folder is not None:
            for candidate in candidates:
                if sheet_folder in candidate.parents:
                    picked = candidate
                    break
        if picked is None and candidates:
            picked = candidates[0]
        if picked is not None:
            results.append(picked.relative_to(ROOT).as_posix())
    return results


def read_readme(workbook) -> dict[str, dict[str, str]]:
    if "README" not in workbook.sheetnames:
        return {}

    sheet = workbook["README"]
    resources = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        product_name = normalize_text(row[0])
        if not product_name:
            continue
        resources[compact_key(product_name)] = {
            "owner": normalize_text(row[1]) if len(row) > 1 else "",
            "specUrl": normalize_text(row[2]) if len(row) > 2 else "",
            "manualUrl": normalize_text(row[3]) if len(row) > 3 else "",
            "softwareUrl": normalize_text(row[4]) if len(row) > 4 else "",
            "notes": normalize_text(row[5]) if len(row) > 5 else "",
        }
    return resources


def normalize_row(row, width: int) -> list[str]:
    values = [normalize_text(value) for value in row]
    if len(values) < width:
        values.extend([""] * (width - len(values)))
    return values[:width]


def workbook_to_catalog() -> dict:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    folder_index, file_index = build_asset_index(ASSET_ROOT)
    readme_resources = read_readme(workbook)

    product_lines = []

    for sheet_name in workbook.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        sheet = workbook[sheet_name]
        if sheet.max_column < 6 or sheet.max_row < 2:
            continue

        rows = [normalize_row(row, sheet.max_column) for row in sheet.iter_rows(values_only=True)]
        header = rows[0]
        data_rows = rows[1:]
        if not any(any(cell for cell in row) for row in data_rows):
            continue

        sheet_folder = best_folder(sheet_name, folder_index)
        items = []
        solution_rows = []

        for idx, row in enumerate(data_rows, start=2):
            name = row[0] if len(row) > 0 else ""
            group = row[1] if len(row) > 1 else ""
            note = row[2] if len(row) > 2 else ""
            image_cell = row[3] if len(row) > 3 else ""
            part_number = row[5] if len(row) > 5 else ""
            description = row[6] if len(row) > 6 else ""
            quantity = row[7] if len(row) > 7 else ""
            solution_raw = row[8] if len(row) > 8 else ""

            if not any([name, group, note, image_cell, part_number, description, quantity, solution_raw]):
                continue

            is_diagram = "system connection diagram" in group.lower() or "系统连接图" in description
            solution_refs = parse_solution_refs(solution_raw)
            assets = resolve_assets(image_cell, sheet_folder, file_index)

            entry = {
                "id": f"{compact_key(sheet_name) or 'sheet'}-{idx}",
                "rowNumber": idx,
                "name": name,
                "group": group or "Ungrouped",
                "note": note,
                "partNumber": part_number.strip(),
                "description": description,
                "quantity": quantity,
                "solutionRefs": solution_refs,
                "images": assets,
                "isDiagram": is_diagram,
            }

            if is_diagram:
                solution_rows.append(entry)
            else:
                if any([entry["partNumber"], entry["name"], entry["description"]]):
                    items.append(entry)

        if not items and not solution_rows:
            continue

        solutions = []
        for order, solution_row in enumerate(solution_rows, start=1):
            refs = solution_row["solutionRefs"] or [f"S{order}"]
            for ref in refs:
                solutions.append(
                    {
                        "id": ref,
                        "label": ref if solution_row["solutionRefs"] else f"方案 {order}",
                        "title": solution_row["description"] or solution_row["name"] or f"{sheet_name} 方案 {order}",
                        "note": solution_row["note"],
                        "images": solution_row["images"],
                    }
                )

        resources = readme_resources.get(compact_key(sheet_name), {})

        product_lines.append(
            {
                "id": compact_key(sheet_name),
                "sheetName": sheet_name,
                "title": sheet_name.strip(),
                "folder": sheet_folder.relative_to(ROOT).as_posix() if sheet_folder else "",
                "headers": header,
                "resources": resources,
                "solutions": solutions,
                "items": items,
            }
        )

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "workbook": WORKBOOK_PATH.name,
        "productLines": product_lines,
    }


def main() -> None:
    catalog = workbook_to_catalog()
    payload = "window.CATALOG = " + json.dumps(catalog, ensure_ascii=False) + ";\n"
    OUTPUT_PATH.write_text(payload, encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
